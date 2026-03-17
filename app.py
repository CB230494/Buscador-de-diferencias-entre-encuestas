# app.py
# -*- coding: utf-8 -*-

import io
import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
import streamlit as st
from pypdf import PdfReader

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    KeepTogether,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Comparador de Formularios PDF", layout="wide")


# =========================================================
# UTILIDADES GENERALES
# =========================================================
def strip_accents(text: str) -> str:
    if text is None:
        return ""
    text = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def normalize_spaces(text: str) -> str:
    if text is None:
        return ""
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def clean_for_compare(text: str) -> str:
    if not text:
        return ""
    text = strip_accents(text.lower())
    text = text.replace("“", '"').replace("”", '"').replace("’", "'")
    text = normalize_spaces(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, clean_for_compare(a), clean_for_compare(b)).ratio()


def safe_text(text: str) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def clean_question_artifacts(text: str) -> str:
    """
    Limpia residuos típicos del PDF:
    - 'pregunta 33.1'
    - '14.1.'
    - referencias pegadas al final
    """
    text = safe_text(text)

    patterns = [
        r"\b[pP]regunta\s+\d{1,2}(?:\.\d+)?\.?$",
        r"\b\d{1,2}(?:\.\d+)?\.?$",
        r"\.\s*\d{1,2}(?:\.\d+)?\.?$",
    ]
    prev = None
    while prev != text:
        prev = text
        for pat in patterns:
            text = re.sub(pat, "", text).strip()

    return safe_text(text)


# =========================================================
# FILTRO DE RUIDO
# =========================================================
def is_noise_line(line: str) -> bool:
    raw = safe_text(line)
    if not raw:
        return True

    low = clean_for_compare(raw)

    noise_prefixes = [
        "nota:",
        "nota.",
        "nota previa",
        "nota condicional",
        "logica condicional",
        "lógica condicional",
        "consentimiento informado",
        "estrategia sembremos seguridad",
        "fin de la encuesta",
        "datos generales de caracter estadistico",
        "datos generales de carácter estadístico",
        "informacion adicional y contacto voluntario",
        "información adicional y contacto voluntario",
        "propuestas ciudadanas para la mejora de la seguridad",
        "confianza policial",
        "delitos",
        "victimizacion",
        "victimización",
        "riesgos sociales y situacionales",
        "contexto territorial y problematicas",
        "contexto territorial y problemáticas",
        "informacion de condiciones institucionales",
        "información de condiciones institucionales",
        "apartado a:",
        "apartado b:",
    ]

    if any(low.startswith(x) for x in noise_prefixes):
        return True

    if re.fullmatch(r"\d+", low):
        return True

    if len(raw) < 90 and raw.isupper() and "?" not in raw:
        return True

    return False


def normalize_option_text(text: str) -> str:
    text = text.strip(" -•\t")
    text = re.sub(r"^[\(\[]?\s*[xX ]?\s*[\)\]]\s*", "", text)
    text = re.sub(r"^[oO]\s+", "", text)
    text = re.sub(r"^[•●▪◦]\s*", "", text)
    text = re.sub(r"^☐\s*", "", text)
    text = re.sub(r"^\(\s*\)\s*", "", text)
    text = re.sub(r"^\-\s*", "", text)
    text = re.sub(r"\s+", " ", text).strip(" .;:")
    return text.strip()


# =========================================================
# EXTRACCIÓN PDF
# =========================================================
def extract_pdf_text(file_obj) -> str:
    try:
        reader = PdfReader(file_obj)
        pages = []
        for page in reader.pages:
            txt = page.extract_text() or ""
            pages.append(txt)
        return normalize_spaces("\n".join(pages))
    except Exception:
        return ""


def detect_tipo(text: str, filename: str = "") -> str:
    base = clean_for_compare((filename or "") + " " + (text[:2000] if text else ""))

    if "encuesta policial" in base or "policial" in base:
        return "Policial"
    if "encuesta comercio" in base or "comercio" in base:
        return "Comercio"
    if "encuesta comunidad" in base or "comunidad" in base:
        return "Comunidad"
    return "Desconocido"


def detect_delegacion(text: str, filename: str = "") -> str:
    source = ((filename or "") + "\n" + (text[:2000] if text else "")).upper()

    m = re.search(r"\bD\s*([0-9]{1,3})\b.*?\b([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ ]{2,})\b", source, re.DOTALL)
    if m:
        d = f"D{m.group(1)}"
        place = m.group(2).strip()
        place = re.sub(r"\s+", " ", place)
        place = re.sub(r"(FORMATO|ENCUESTA|COMUNIDAD|COMERCIO|POLICIAL)$", "", place).strip()
        if place:
            return f"{d} - {place.title()}"

    m2 = re.search(r"\bD\s*([0-9]{1,3})\b", source)
    if m2:
        return f"D{m2.group(1)}"

    return "No identificada"


# =========================================================
# PARSER DE PREGUNTAS Y OPCIONES
# =========================================================
QUESTION_RE = re.compile(r"^\s*(\d{1,2}(?:\.\d+)?)[\.\-–]?\s*(.+)$")

OPTION_HINT_WORDS = [
    "si",
    "no",
    "sí",
    "muy inseguro",
    "inseguro",
    "seguro",
    "muy seguro",
    "nunca",
    "casi nunca",
    "todos los dias",
    "todos los días",
    "varias veces",
    "una vez",
    "otro",
    "otra",
    "no aplica",
    "desconocido",
    "arma",
    "hurto",
    "robo",
    "asalto",
    "estafa",
    "extorsion",
    "extorsión",
]


def split_lines_safely(text: str) -> list[str]:
    lines = []

    for raw in text.split("\n"):
        part = raw.strip()
        if not part:
            continue

        part = re.sub(r"(?<!^)\s+(\d{1,2}(?:\.\d+)?[\.\-–])\s*", r"\n\1 ", part)

        for p in part.split("\n"):
            p = p.strip()
            if p:
                lines.append(p)

    return lines


def is_question_line(line: str) -> bool:
    if is_noise_line(line):
        return False

    m = QUESTION_RE.match(line)
    if not m:
        return False

    q_text = m.group(2).strip()
    return len(q_text) >= 3


def looks_like_option(line: str) -> bool:
    if is_noise_line(line):
        return False

    raw = line.strip()
    low = clean_for_compare(raw)

    if raw.startswith(("☐", "•", "●", "▪", "◦")):
        return True
    if re.match(r"^\(\s*\)\s*", raw):
        return True
    if re.match(r"^[oO]\s+", raw):
        return True

    if len(raw) <= 140 and any(word in low for word in OPTION_HINT_WORDS):
        return True

    return False


def parse_questions(text: str) -> list[dict]:
    lines = split_lines_safely(text)

    questions = []
    current = None

    for line in lines:
        if is_question_line(line):
            m = QUESTION_RE.match(line)
            q_num = m.group(1).strip()
            q_text = m.group(2).strip()

            current = {
                "num": q_num,
                "question": clean_question_artifacts(q_text),
                "options": []
            }
            questions.append(current)
            continue

        if current is None:
            continue

        if is_noise_line(line):
            continue

        if looks_like_option(line):
            opt = normalize_option_text(line)
            if opt:
                current["options"].append(opt)
            continue

        low = clean_for_compare(line)
        if not low.startswith(("nota", "logica", "lógica", "si la respuesta", "en caso de", "al continuar")):
            if len(line) < 220:
                if not re.fullmatch(r"\d{1,2}(?:\.\d+)?\.?", line.strip()):
                    current["question"] = clean_question_artifacts(
                        safe_text(current["question"] + " " + line)
                    )

    cleaned = []
    for q in questions:
        q_text = clean_question_artifacts(q["question"])
        opts = []
        seen = set()

        for op in q["options"]:
            op2 = safe_text(op)
            op_key = clean_for_compare(op2)
            if op2 and op_key not in seen:
                seen.add(op_key)
                opts.append(op2)

        if q_text:
            cleaned.append({
                "num": q["num"],
                "question": q_text,
                "options": opts
            })

    return cleaned


# =========================================================
# COMPARACIÓN
# =========================================================
def build_question_map(questions: list[dict]) -> dict:
    return {q["num"]: q for q in questions}


def compare_options(orig_opts: list[str], new_opts: list[str]) -> list[dict]:
    changes = []

    orig_norm = {clean_for_compare(x): x for x in orig_opts}
    new_norm = {clean_for_compare(x): x for x in new_opts}

    orig_keys = set(orig_norm.keys())
    new_keys = set(new_norm.keys())

    added = [new_norm[k] for k in sorted(new_keys - orig_keys)]
    removed = [orig_norm[k] for k in sorted(orig_keys - new_keys)]

    used_added = set()
    used_removed = set()

    for i, rem in enumerate(removed):
        best_j = None
        best_score = 0.0
        for j, add in enumerate(added):
            if j in used_added:
                continue
            sc = similarity(rem, add)
            if sc > best_score:
                best_score = sc
                best_j = j

        if best_j is not None and best_score >= 0.60:
            changes.append({
                "type": "opcion_modificada",
                "antes": rem,
                "despues": added[best_j]
            })
            used_removed.add(i)
            used_added.add(best_j)

    for i, rem in enumerate(removed):
        if i not in used_removed:
            changes.append({
                "type": "opcion_eliminada",
                "texto": rem
            })

    for j, add in enumerate(added):
        if j not in used_added:
            changes.append({
                "type": "opcion_agregada",
                "texto": add
            })

    return changes


def compare_questions(orig_questions: list[dict], new_questions: list[dict]) -> list[dict]:
    changes = []

    orig_map = build_question_map(orig_questions)
    new_map = build_question_map(new_questions)

    all_nums = sorted(
        set(orig_map.keys()) | set(new_map.keys()),
        key=lambda x: [int(p) if p.isdigit() else p for p in re.split(r"(\d+)", x)]
    )

    for num in all_nums:
        oq = orig_map.get(num)
        nq = new_map.get(num)

        if oq and not nq:
            changes.append({
                "question_num": num,
                "question_label": f"Pregunta {num}",
                "change_kind": "pregunta_eliminada",
                "detail": f"La pregunta {num} existe en el original pero no aparece en el modificado."
            })
            continue

        if nq and not oq:
            changes.append({
                "question_num": num,
                "question_label": f"Pregunta {num}",
                "change_kind": "pregunta_agregada",
                "detail": f"Se agregó la pregunta {num}: {nq['question']}"
            })
            continue

        q_changes = []

        q_sim = similarity(oq["question"], nq["question"])
        if q_sim < 0.97:
            q_changes.append({
                "type": "texto_pregunta_modificado",
                "antes": clean_question_artifacts(oq["question"]),
                "despues": clean_question_artifacts(nq["question"])
            })

        opt_changes = compare_options(oq["options"], nq["options"])
        q_changes.extend(opt_changes)

        if q_changes:
            changes.append({
                "question_num": num,
                "question_label": f"Pregunta {num}",
                "change_kind": "detalle",
                "question_original": oq["question"],
                "question_new": nq["question"],
                "changes": q_changes
            })

    return changes


# =========================================================
# PDF
# =========================================================
def build_detailed_pdf(report_rows: list[dict]) -> bytes:
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=28,
        leftMargin=28,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    style_title = ParagraphStyle(
        "title_custom",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=1,
        spaceAfter=10,
    )

    style_intro = ParagraphStyle(
        "intro_custom",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=12,
        alignment=0,
        spaceAfter=10,
    )

    style_h2 = ParagraphStyle(
        "h2_custom",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
        spaceBefore=8,
        spaceAfter=6,
    )

    style_h3 = ParagraphStyle(
        "h3_custom",
        parent=styles["Heading3"],
        fontName="Helvetica-BoldOblique",
        fontSize=11,
        leading=13,
        spaceBefore=4,
        spaceAfter=4,
    )

    style_meta = ParagraphStyle(
        "meta_custom",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=11,
        spaceAfter=6,
        wordWrap="CJK",
    )

    style_box = ParagraphStyle(
        "box_custom",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        leftIndent=8,
        rightIndent=4,
        spaceBefore=1,
        spaceAfter=1,
        wordWrap="CJK",
    )

    story = []

    story.append(Paragraph("Reporte comparativo de preguntas y opciones", style_title))
    story.append(Paragraph(
        "Solo se incluyen cambios sustantivos en preguntas y respuestas. "
        "No se incorporan notas, leyendas, instrucciones ni texto introductorio.",
        style_intro
    ))
    story.append(Spacer(1, 8))

    summary_data = [["Archivo", "Tipo", "Delegación", "Cambios detectados"]]

    for row in report_rows:
        summary_data.append([
            Paragraph(safe_text(row["archivo"]), style_meta),
            Paragraph(safe_text(row["tipo"]), style_meta),
            Paragraph(safe_text(row["delegacion"]), style_meta),
            Paragraph(str(row["total_cambios"]), style_meta),
        ])

    summary_table = Table(
        summary_data,
        repeatRows=1,
        colWidths=[245, 90, 130, 75]
    )

    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#EDEDED")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))

    story.append(summary_table)
    story.append(Spacer(1, 14))

    for idx, row in enumerate(report_rows, start=1):
        story.append(Paragraph(f"{idx}. {safe_text(row['archivo'])}", style_h2))
        story.append(Paragraph(
            f"<b>Tipo:</b> {safe_text(row['tipo'])} &nbsp;&nbsp;&nbsp; "
            f"<b>Delegación:</b> {safe_text(row['delegacion'])} &nbsp;&nbsp;&nbsp; "
            f"<b>Total de cambios:</b> {row['total_cambios']}",
            style_meta
        ))
        story.append(Spacer(1, 4))

        if row["error"]:
            story.append(Paragraph(f"<b>Error:</b> {safe_text(row['error'])}", style_box))
            story.append(Spacer(1, 10))
            continue

        if not row["cambios"]:
            story.append(Paragraph(
                "No se detectaron cambios sustantivos en preguntas u opciones.",
                style_box
            ))
            story.append(Spacer(1, 10))
            continue

        for item in row["cambios"]:
            bloque_pregunta = [
                Paragraph(safe_text(item["question_label"]), style_h3)
            ]

            if item["change_kind"] in ("pregunta_agregada", "pregunta_eliminada"):
                bloque_pregunta.append(Paragraph(safe_text(item["detail"]), style_box))
                bloque_pregunta.append(Spacer(1, 5))
                story.append(KeepTogether(bloque_pregunta))
                continue

            for ch in item["changes"]:
                if ch["type"] == "texto_pregunta_modificado":
                    bloque_pregunta.append(Paragraph("<b>Texto modificado</b>", style_box))
                    bloque_pregunta.append(
                        Paragraph(f"<b>Original:</b> {safe_text(ch['antes'])}", style_box)
                    )
                    bloque_pregunta.append(
                        Paragraph(f"<b>Nuevo:</b> {safe_text(ch['despues'])}", style_box)
                    )

                elif ch["type"] == "opcion_agregada":
                    bloque_pregunta.append(
                        Paragraph(f"<b>Opción agregada:</b> {safe_text(ch['texto'])}", style_box)
                    )

                elif ch["type"] == "opcion_eliminada":
                    bloque_pregunta.append(
                        Paragraph(f"<b>Opción eliminada:</b> {safe_text(ch['texto'])}", style_box)
                    )

                elif ch["type"] == "opcion_modificada":
                    sub_bloque = [
                        Paragraph("<b>Opción modificada</b>", style_box),
                        Paragraph(f"<b>Antes:</b> {safe_text(ch['antes'])}", style_box),
                        Paragraph(f"<b>Después:</b> {safe_text(ch['despues'])}", style_box),
                    ]
                    bloque_pregunta.extend(sub_bloque)

            bloque_pregunta.append(Spacer(1, 6))
            story.append(KeepTogether(bloque_pregunta))

        story.append(Spacer(1, 10))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


# =========================================================
# EXCEL
# =========================================================
def flatten_report_rows(report_rows: list[dict]) -> tuple[pd.DataFrame, pd.DataFrame]:
    resumen_rows = []
    detalle_rows = []

    for row in report_rows:
        resumen_rows.append({
            "Archivo": Path(row["archivo"]).name,
            "Tipo": safe_text(row["tipo"]),
            "Delegación": safe_text(row["delegacion"]),
            "Cambios detectados": row["total_cambios"],
            "Estado": "Error" if row["error"] else ("Sin cambios" if row["total_cambios"] == 0 else "Con cambios"),
            "Observación": safe_text(row["error"]) if row["error"] else (
                "No se detectaron cambios sustantivos en preguntas u opciones."
                if row["total_cambios"] == 0 else
                "Se detectaron cambios en preguntas y/o opciones."
            )
        })

        if row["error"]:
            detalle_rows.append({
                "Archivo": Path(row["archivo"]).name,
                "Tipo": safe_text(row["tipo"]),
                "Delegación": safe_text(row["delegacion"]),
                "Pregunta Nº": "",
                "Etiqueta": "",
                "Categoría de cambio": "Error",
                "Pregunta original": "",
                "Pregunta nueva": "",
                "Valor original": "",
                "Valor nuevo": "",
                "Detalle / nota": safe_text(row["error"])
            })
            continue

        if not row["cambios"]:
            detalle_rows.append({
                "Archivo": Path(row["archivo"]).name,
                "Tipo": safe_text(row["tipo"]),
                "Delegación": safe_text(row["delegacion"]),
                "Pregunta Nº": "",
                "Etiqueta": "",
                "Categoría de cambio": "Sin cambios",
                "Pregunta original": "",
                "Pregunta nueva": "",
                "Valor original": "",
                "Valor nuevo": "",
                "Detalle / nota": "No se detectaron cambios sustantivos en preguntas u opciones."
            })
            continue

        for item in row["cambios"]:
            if item["change_kind"] == "pregunta_agregada":
                detalle_rows.append({
                    "Archivo": Path(row["archivo"]).name,
                    "Tipo": safe_text(row["tipo"]),
                    "Delegación": safe_text(row["delegacion"]),
                    "Pregunta Nº": safe_text(item.get("question_num", "")),
                    "Etiqueta": safe_text(item.get("question_label", "")),
                    "Categoría de cambio": "Pregunta agregada",
                    "Pregunta original": "",
                    "Pregunta nueva": "",
                    "Valor original": "",
                    "Valor nuevo": "",
                    "Detalle / nota": safe_text(item.get("detail", ""))
                })
                continue

            if item["change_kind"] == "pregunta_eliminada":
                detalle_rows.append({
                    "Archivo": Path(row["archivo"]).name,
                    "Tipo": safe_text(row["tipo"]),
                    "Delegación": safe_text(row["delegacion"]),
                    "Pregunta Nº": safe_text(item.get("question_num", "")),
                    "Etiqueta": safe_text(item.get("question_label", "")),
                    "Categoría de cambio": "Pregunta eliminada",
                    "Pregunta original": "",
                    "Pregunta nueva": "",
                    "Valor original": "",
                    "Valor nuevo": "",
                    "Detalle / nota": safe_text(item.get("detail", ""))
                })
                continue

            for ch in item.get("changes", []):
                categoria = ""
                val_original = ""
                val_nuevo = ""
                detalle = ""

                if ch["type"] == "texto_pregunta_modificado":
                    categoria = "Texto de pregunta modificado"
                    val_original = safe_text(ch.get("antes", ""))
                    val_nuevo = safe_text(ch.get("despues", ""))
                    detalle = "Se modificó el texto de la pregunta."

                elif ch["type"] == "opcion_agregada":
                    categoria = "Opción agregada"
                    val_nuevo = safe_text(ch.get("texto", ""))
                    detalle = "Se agregó una nueva opción de respuesta."

                elif ch["type"] == "opcion_eliminada":
                    categoria = "Opción eliminada"
                    val_original = safe_text(ch.get("texto", ""))
                    detalle = "Se eliminó una opción existente."

                elif ch["type"] == "opcion_modificada":
                    categoria = "Opción modificada"
                    val_original = safe_text(ch.get("antes", ""))
                    val_nuevo = safe_text(ch.get("despues", ""))
                    detalle = "Se modificó una opción de respuesta."

                detalle_rows.append({
                    "Archivo": Path(row["archivo"]).name,
                    "Tipo": safe_text(row["tipo"]),
                    "Delegación": safe_text(row["delegacion"]),
                    "Pregunta Nº": safe_text(item.get("question_num", "")),
                    "Etiqueta": safe_text(item.get("question_label", "")),
                    "Categoría de cambio": categoria,
                    "Pregunta original": safe_text(item.get("question_original", "")),
                    "Pregunta nueva": safe_text(item.get("question_new", "")),
                    "Valor original": val_original,
                    "Valor nuevo": val_nuevo,
                    "Detalle / nota": detalle
                })

    resumen_df = pd.DataFrame(resumen_rows)
    detalle_df = pd.DataFrame(detalle_rows)

    return resumen_df, detalle_df


def auto_adjust_width(ws, min_width=12, max_width=45):
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            value = value.replace("\n", " ")
            max_len = max(max_len, len(value))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_sheet(ws, title: str, header_row: int, data_start_row: int):
    blue_fill = PatternFill("solid", fgColor="1F4E79")
    light_blue_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    dark_font = Font(color="000000", bold=True)
    thin = Side(style="thin", color="B7B7B7")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.fill = blue_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = light_blue_fill
        cell.font = dark_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(data_start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = f"A{data_start_row}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def apply_detail_row_colors(ws, header_row: int, data_start_row: int):
    fills = {
        "Texto de pregunta modificado": PatternFill("solid", fgColor="FFF2CC"),
        "Opción agregada": PatternFill("solid", fgColor="E2F0D9"),
        "Opción eliminada": PatternFill("solid", fgColor="FCE4D6"),
        "Opción modificada": PatternFill("solid", fgColor="D9EAD3"),
        "Pregunta agregada": PatternFill("solid", fgColor="DDEBF7"),
        "Pregunta eliminada": PatternFill("solid", fgColor="F4CCCC"),
        "Sin cambios": PatternFill("solid", fgColor="EDEDED"),
        "Error": PatternFill("solid", fgColor="F4CCCC"),
    }

    cat_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "Categoría de cambio":
            cat_col = col
            break

    if cat_col is None:
        return

    for row in range(data_start_row, ws.max_row + 1):
        categoria = ws.cell(row=row, column=cat_col).value
        fill = fills.get(str(categoria).strip()) if categoria is not None else None
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


def apply_summary_row_colors(ws, header_row: int, data_start_row: int):
    fills = {
        "Con cambios": PatternFill("solid", fgColor="FFF2CC"),
        "Sin cambios": PatternFill("solid", fgColor="E2F0D9"),
        "Error": PatternFill("solid", fgColor="F4CCCC"),
    }

    estado_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "Estado":
            estado_col = col
            break

    if estado_col is None:
        return

    for row in range(data_start_row, ws.max_row + 1):
        estado = ws.cell(row=row, column=estado_col).value
        fill = fills.get(str(estado).strip()) if estado is not None else None
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill


def build_excel_report(report_rows: list[dict]) -> bytes:
    resumen_df, detalle_df = flatten_report_rows(report_rows)

    wb = Workbook()

    # -----------------------------------------------------
    # Hoja Resumen
    # -----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Resumen General"

    resumen_title = "Reporte comparativo de preguntas y opciones - Resumen General"
    headers_1 = list(resumen_df.columns)

    start_header_row_1 = 4
    start_data_row_1 = 5

    ws1["A2"] = "Descripción:"
    ws1["B2"] = "Resumen por archivo comparado, indicando tipo, delegación, cantidad de cambios y observación general."

    for idx, col_name in enumerate(headers_1, start=1):
        ws1.cell(row=start_header_row_1, column=idx, value=col_name)

    for r_idx, row in enumerate(resumen_df.itertuples(index=False), start=start_data_row_1):
        for c_idx, value in enumerate(row, start=1):
            ws1.cell(row=r_idx, column=c_idx, value=value)

    style_sheet(ws1, resumen_title, start_header_row_1, start_data_row_1)
    apply_summary_row_colors(ws1, start_header_row_1, start_data_row_1)
    auto_adjust_width(ws1)

    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 36

    # -----------------------------------------------------
    # Hoja Detalle
    # -----------------------------------------------------
    ws2 = wb.create_sheet("Detalle de Cambios")

    detalle_title = "Reporte comparativo de preguntas y opciones - Detalle de cambios"
    headers_2 = list(detalle_df.columns)

    start_header_row_2 = 4
    start_data_row_2 = 5

    ws2["A2"] = "Descripción:"
    ws2["B2"] = (
        "Cada fila representa un cambio específico. "
        "No se acumulan varios cambios dentro de una sola celda para mantener el orden y facilitar filtros."
    )

    for idx, col_name in enumerate(headers_2, start=1):
        ws2.cell(row=start_header_row_2, column=idx, value=col_name)

    for r_idx, row in enumerate(detalle_df.itertuples(index=False), start=start_data_row_2):
        for c_idx, value in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    style_sheet(ws2, detalle_title, start_header_row_2, start_data_row_2)
    apply_detail_row_colors(ws2, start_header_row_2, start_data_row_2)
    auto_adjust_width(ws2)

    ws2.row_dimensions[1].height = 24
    ws2.row_dimensions[2].height = 42

    # Anchos específicos útiles
    custom_widths = {
        "A": 28,
        "B": 16,
        "C": 24,
        "D": 12,
        "E": 16,
        "F": 28,
        "G": 45,
        "H": 45,
        "I": 38,
        "J": 38,
        "K": 36,
    }
    for col, width in custom_widths.items():
        ws2.column_dimensions[col].width = width

    # -----------------------------------------------------
    # Hoja Notas
    # -----------------------------------------------------
    ws3 = wb.create_sheet("Notas")

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "Notas del reporte"
    ws3["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws3["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")

    notes_headers = ["Elemento", "Significado", "Color de referencia", "Uso recomendado"]
    notes_data = [
        ["Con cambios", "El archivo presenta cambios detectados.", "Amarillo suave", "Revisar detalle."],
        ["Sin cambios", "No se detectaron cambios sustantivos.", "Verde suave", "Archivo estable."],
        ["Error", "No se pudo comparar correctamente o no se encontró original compatible.", "Rojo suave", "Verificar archivo original/modificado."],
        ["Texto de pregunta modificado", "Se alteró el texto principal de la pregunta.", "Amarillo", "Revisar redacción."],
        ["Opción agregada", "Se añadió una nueva opción de respuesta.", "Verde", "Revisar si procede."],
        ["Opción eliminada", "Se removió una opción previa.", "Naranja/rojo suave", "Verificar impacto."],
        ["Opción modificada", "Una opción cambió respecto al original.", "Verde claro", "Comparar precisión del cambio."],
        ["Pregunta agregada/eliminada", "Se añadió o quitó una pregunta completa.", "Azul / Rojo suave", "Validar estructura del formulario."],
    ]

    for idx, h in enumerate(notes_headers, start=1):
        ws3.cell(row=3, column=idx, value=h)

    for r_idx, row in enumerate(notes_data, start=4):
        for c_idx, value in enumerate(row, start=1):
            ws3.cell(row=r_idx, column=c_idx, value=value)

    style_sheet(ws3, "Guía de lectura del Excel", 3, 4)
    auto_adjust_width(ws3, min_width=16, max_width=42)

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 46
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 28

    # -----------------------------------------------------
    # Guardar
    # -----------------------------------------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =========================================================
# LÓGICA
# =========================================================
def build_reference_map(files: list) -> dict:
    ref_map = {}

    for f in files:
        text = extract_pdf_text(f)
        tipo = detect_tipo(text, f.name)
        delegacion = detect_delegacion(text, f.name)
        questions = parse_questions(text)

        key = tipo.lower()
        ref_map[key] = {
            "filename": f.name,
            "tipo": tipo,
            "delegacion": delegacion,
            "text": text,
            "questions": questions,
        }

    return ref_map


# =========================================================
# INTERFAZ
# =========================================================
st.title("Comparador de Formularios PDF")
st.write(
    "Cargue primero los **PDF originales** y luego los **PDF con cambios**. "
    "La comparación se realiza solo sobre **preguntas y opciones de respuesta**."
)

col1, col2 = st.columns(2)

with col1:
    st.subheader("1. PDF originales")
    original_files = st.file_uploader(
        "Cargue los originales",
        type=["pdf"],
        accept_multiple_files=True,
        key="orig"
    )

with col2:
    st.subheader("2. PDF modificados")
    modified_files = st.file_uploader(
        "Cargue los modificados",
        type=["pdf"],
        accept_multiple_files=True,
        key="mod"
    )

compare_btn = st.button("Comparar formularios", type="primary")

if compare_btn:
    if not original_files or not modified_files:
        st.warning("Debe cargar los PDFs originales y los PDFs modificados.")
        st.stop()

    with st.spinner("Procesando PDFs y comparando preguntas/opciones..."):
        originals = build_reference_map(original_files)
        report_rows = []

        for f in modified_files:
            text = extract_pdf_text(f)
            tipo = detect_tipo(text, f.name)
            delegacion = detect_delegacion(text, f.name)
            questions = parse_questions(text)

            original = originals.get(tipo.lower())

            if not original:
                report_rows.append({
                    "archivo": f.name,
                    "tipo": tipo,
                    "delegacion": delegacion,
                    "total_cambios": 0,
                    "cambios": [],
                    "error": "No se encontró original compatible."
                })
                continue

            cambios = compare_questions(original["questions"], questions)

            report_rows.append({
                "archivo": f.name,
                "tipo": tipo,
                "delegacion": delegacion,
                "total_cambios": len(cambios),
                "cambios": cambios,
                "error": None
            })

    st.subheader("Resumen general")

    summary_df = pd.DataFrame([
        {
            "Archivo": Path(r["archivo"]).name,
            "Tipo": r["tipo"],
            "Delegación": r["delegacion"],
            "Cambios detectados": r["total_cambios"]
        }
        for r in report_rows
    ])

    st.dataframe(summary_df, use_container_width=True)

    st.subheader("Detalle de cambios")

    for row in report_rows:
        st.markdown("---")
        st.markdown(f"### {row['archivo']}")
        st.write(f"**Tipo:** {row['tipo']}")
        st.write(f"**Delegación:** {row['delegacion']}")
        st.write(f"**Total de cambios:** {row['total_cambios']}")

        if row["error"]:
            st.error(row["error"])
            continue

        if not row["cambios"]:
            st.success("No se detectaron cambios sustantivos en preguntas u opciones.")
            continue

        for item in row["cambios"]:
            st.markdown(f"#### {item['question_label']}")

            if item["change_kind"] in ("pregunta_agregada", "pregunta_eliminada"):
                st.warning(item["detail"])
                continue

            for ch in item["changes"]:
                if ch["type"] == "texto_pregunta_modificado":
                    st.info("Texto de la pregunta modificado")
                    st.write(f"**Original:** {ch['antes']}")
                    st.write(f"**Nuevo:** {ch['despues']}")

                elif ch["type"] == "opcion_agregada":
                    st.success(f"**Opción agregada:** {ch['texto']}")

                elif ch["type"] == "opcion_eliminada":
                    st.error(f"**Opción eliminada:** {ch['texto']}")

                elif ch["type"] == "opcion_modificada":
                    st.warning("Opción modificada")
                    st.write(f"**Antes:** {ch['antes']}")
                    st.write(f"**Después:** {ch['despues']}")

    pdf_bytes = build_detailed_pdf(report_rows)
    excel_bytes = build_excel_report(report_rows)

    col_pdf, col_excel = st.columns(2)

    with col_pdf:
        st.download_button(
            "Descargar reporte PDF detallado",
            data=pdf_bytes,
            file_name="reporte_comparativo_preguntas_opciones.pdf",
            mime="application/pdf",
            use_container_width=True
        )

    with col_excel:
        st.download_button(
            "Descargar reporte Excel detallado",
            data=excel_bytes,
            file_name="reporte_comparativo_preguntas_opciones.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
